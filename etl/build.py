"""Build read-only h1b_data.db from DOL LCA + USCIS Employer Data Hub files."""

from __future__ import annotations

import argparse
import csv
import json
import sqlite3
import statistics
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from etl.canonicalize import canonicalize
from etl.column_maps import resolve_dol_columns, resolve_uscis_columns
from etl.sources import fiscal_year_from_dol_filename, last_n_complete_fiscal_years

ANNUAL_HOURS = 2080
CERTIFIED_STATUSES = frozenset({"CERTIFIED", "CERTIFIED-EXPIRED"})


@dataclass
class YearBucket:
    certified: int = 0
    salaries: list[float] = field(default_factory=list)
    titles: Counter[str] = field(default_factory=Counter)
    new_approvals: int = 0
    new_denials: int = 0
    # None = source vintage has no Change of Employer breakout (NULL in DB)
    transfer_approvals: int | None = None
    transfer_denials: int | None = None


def annualize_wage(amount: float, unit: str) -> float | None:
    if amount <= 0:
        return None
    u = (unit or "").strip().lower()
    if u in ("year", "yr"):
        return amount
    if u == "hour":
        return amount * ANNUAL_HOURS
    if u == "month":
        return amount * 12
    if u == "week":
        return amount * 52
    if u in ("bi-weekly", "biweekly"):
        return amount * 26
    return None


def _col_index(header: list[str], name: str) -> int | None:
    try:
        return header.index(name)
    except ValueError:
        return None


def ingest_dol_xlsx(path: Path, buckets: dict[tuple[str, int], YearBucket]) -> int:
    """Stream one DOL quarterly xlsx; return fiscal year inferred from filename."""
    fy = fiscal_year_from_dol_filename(path.name)
    if fy is None:
        raise ValueError(f"Cannot infer fiscal year from DOL filename: {path.name}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        wb.close()
        return fy

    header = [str(c).strip() if c is not None else "" for c in header_row]
    cols = resolve_dol_columns(tuple(header))
    if cols is None:
        wb.close()
        raise ValueError(f"No recognized DOL columns in {path}")

    idx_status = _col_index(header, cols.case_status)
    idx_employer = _col_index(header, cols.employer_name)
    idx_title = _col_index(header, cols.job_title)
    idx_wage = _col_index(header, cols.wage_from)
    idx_unit = _col_index(header, cols.wage_unit)
    if None in (idx_status, idx_employer, idx_title, idx_wage, idx_unit):
        wb.close()
        raise ValueError(f"Missing required DOL columns in {path}")

    for row in rows:
        if not row or len(row) <= max(idx_status, idx_employer, idx_title, idx_wage, idx_unit):
            continue
        status = str(row[idx_status] or "").strip().upper()
        if status not in CERTIFIED_STATUSES:
            continue
        filed = str(row[idx_employer] or "").strip()
        if not filed:
            continue
        canonical = canonicalize(filed)
        if not canonical:
            continue
        key = (canonical, fy)
        bucket = buckets[key]
        bucket.certified += 1
        title = str(row[idx_title] or "").strip()
        if title:
            bucket.titles[title.upper()] += 1
        try:
            wage_raw = row[idx_wage]
            wage = float(wage_raw) if wage_raw is not None else 0.0
        except (TypeError, ValueError):
            wage = 0.0
        unit = str(row[idx_unit] or "")
        annual = annualize_wage(wage, unit)
        if annual is not None:
            bucket.salaries.append(annual)

    wb.close()
    return fy


def _open_uscis(path: Path):
    """Open a USCIS CSV, detecting UTF-16 (real Data Hub export, tab-delimited)
    vs UTF-8 (synthetic fixture, comma-delimited) from the byte-order mark."""
    with path.open("rb") as fh:
        bom = fh.read(2)
    if bom in (b"\xff\xfe", b"\xfe\xff"):
        return path.open(newline="", encoding="utf-16"), "\t"
    return path.open(newline="", encoding="utf-8-sig"), ","


def _uscis_indices(header: tuple[str, ...], cols):
    """Column indices, tolerant of trailing/leading whitespace in headers
    (the real export has 'Fiscal Year   ')."""
    stripped = [h.strip() for h in header]

    def idx(name: str) -> int:
        return stripped.index(name)

    return (
        idx(cols.employer),
        idx(cols.fiscal_year),
        [idx(c) for c in cols.new_approval_columns],
        [idx(c) for c in cols.new_denial_columns],
        [idx(c) for c in cols.transfer_approval_columns],
        [idx(c) for c in cols.transfer_denial_columns],
    )


def _sum_cells(row: list[str], indices: list[int]) -> int:
    total = 0
    for i in indices:
        total += int(str(row[i]).replace(",", "").strip() or 0)
    return total


def _accumulate_uscis(
    bucket: YearBucket,
    row: list,
    idx_new_app: list[int],
    idx_new_den: list[int],
    idx_tr_app: list[int],
    idx_tr_den: list[int],
) -> None:
    bucket.new_approvals += _sum_cells(row, idx_new_app)
    bucket.new_denials += _sum_cells(row, idx_new_den)
    if idx_tr_app:  # breakout available in this vintage
        bucket.transfer_approvals = (bucket.transfer_approvals or 0) + _sum_cells(row, idx_tr_app)
        bucket.transfer_denials = (bucket.transfer_denials or 0) + _sum_cells(row, idx_tr_den)


def ingest_uscis_csv(
    path: Path,
    buckets: dict[tuple[str, int], YearBucket],
    filed_names: dict[str, set[str]] | None = None,
) -> None:
    fh, delim = _open_uscis(path)
    with fh:
        reader = csv.reader(fh, delimiter=delim)
        header = tuple(next(reader, []))
        cols = resolve_uscis_columns(header)
        idx_emp, idx_fy, idx_new_app, idx_new_den, idx_tr_app, idx_tr_den = _uscis_indices(
            header, cols
        )
        max_idx = max([idx_emp, idx_fy, *idx_new_app, *idx_new_den, *idx_tr_app, *idx_tr_den])

        for row in reader:
            if len(row) <= max_idx:
                continue
            filed = row[idx_emp].strip()
            if not filed:
                continue
            try:
                fy = int(str(row[idx_fy]).strip())
            except ValueError:
                continue
            canonical = canonicalize(filed)
            bucket = buckets.setdefault((canonical, fy), YearBucket())
            try:
                _accumulate_uscis(bucket, row, idx_new_app, idx_new_den, idx_tr_app, idx_tr_den)
            except ValueError:
                continue
            if filed_names is not None:
                filed_names[filed.upper()].add(canonical)


def _salary_stats(salaries: list[float]) -> tuple[float | None, float | None, float | None]:
    if not salaries:
        return None, None, None
    return statistics.median(salaries), min(salaries), max(salaries)


def _top_titles(counter: Counter[str], n: int = 5) -> list[str]:
    return [t for t, _ in counter.most_common(n)]


def write_database(
    db_path: Path,
    buckets: dict[tuple[str, int], YearBucket],
    filed_names: dict[str, set[str]],
) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    if db_path.exists():
        db_path.unlink()

    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=OFF")
    conn.executescript(
        """
        CREATE TABLE employers (
            canonical_employer TEXT PRIMARY KEY
        );
        CREATE TABLE filed_names (
            filed_name TEXT NOT NULL,
            canonical_employer TEXT NOT NULL,
            PRIMARY KEY (filed_name, canonical_employer)
        );
        CREATE TABLE aggregates (
            canonical_employer TEXT NOT NULL,
            fiscal_year INTEGER NOT NULL,
            certified_count INTEGER NOT NULL DEFAULT 0,
            salary_median REAL,
            salary_min REAL,
            salary_max REAL,
            top_titles TEXT NOT NULL DEFAULT '[]',
            uscis_new_approvals INTEGER NOT NULL DEFAULT 0,
            uscis_new_denials INTEGER NOT NULL DEFAULT 0,
            uscis_transfer_approvals INTEGER,
            uscis_transfer_denials INTEGER,
            PRIMARY KEY (canonical_employer, fiscal_year)
        );
        CREATE VIRTUAL TABLE employer_search USING fts5(
            canonical_employer,
            filed_name,
            content='',
            tokenize='unicode61 remove_diacritics 1'
        );
        CREATE TABLE meta (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        """
    )

    employers = {canon for canon, _ in buckets}
    for emp in sorted(employers):
        conn.execute("INSERT INTO employers VALUES (?)", (emp,))

    for filed, canon_set in filed_names.items():
        for canon in canon_set:
            conn.execute(
                "INSERT OR IGNORE INTO filed_names VALUES (?, ?)",
                (filed.upper(), canon),
            )

    for (canon, fy), bucket in sorted(buckets.items()):
        med, smin, smax = _salary_stats(bucket.salaries)
        conn.execute(
            """
            INSERT INTO aggregates (
                canonical_employer, fiscal_year, certified_count,
                salary_median, salary_min, salary_max, top_titles,
                uscis_new_approvals, uscis_new_denials,
                uscis_transfer_approvals, uscis_transfer_denials
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                canon,
                fy,
                bucket.certified,
                med,
                smin,
                smax,
                json.dumps(_top_titles(bucket.titles)),
                bucket.new_approvals,
                bucket.new_denials,
                bucket.transfer_approvals,
                bucket.transfer_denials,
            ),
        )

    # Populate FTS with canonical + filed names for fuzzy lookup
    for emp in sorted(employers):
        conn.execute(
            "INSERT INTO employer_search(canonical_employer, filed_name) VALUES (?, ?)",
            (emp, emp),
        )
    for row in conn.execute("SELECT filed_name, canonical_employer FROM filed_names"):
        conn.execute(
            "INSERT INTO employer_search(canonical_employer, filed_name) VALUES (?, ?)",
            (row[1], row[0]),
        )

    conn.execute(
        """
        INSERT INTO meta (key, value) VALUES ('latest_complete_fy', ?),
               ('built_fiscal_years', ?)
        """,
        (
            str(last_n_complete_fiscal_years(1)[0]),
            json.dumps(sorted({fy for _, fy in buckets})),
        ),
    )
    conn.commit()
    conn.close()


def build_from_paths(
    dol_paths: list[Path],
    uscis_paths: list[Path],
    output: Path,
) -> None:
    buckets: dict[tuple[str, int], YearBucket] = defaultdict(YearBucket)
    filed_names: dict[str, set[str]] = defaultdict(set)

    for dol_path in dol_paths:
        fy = ingest_dol_xlsx(dol_path, buckets)
        # Track filed names from aggregates keys for this FY
        for (canon, year), bucket in list(buckets.items()):
            if year == fy and bucket.certified:
                filed_names[canon].add(canon)

    # Re-scan DOL for filed name variants (second pass on smaller fixture files)
    for dol_path in dol_paths:
        fy = fiscal_year_from_dol_filename(dol_path.name)
        if fy is None:
            continue
        wb = load_workbook(dol_path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            wb.close()
            continue
        header = [str(c).strip() if c is not None else "" for c in header_row]
        cols = resolve_dol_columns(tuple(header))
        if cols is None:
            wb.close()
            continue
        idx_status = header.index(cols.case_status)
        idx_employer = header.index(cols.employer_name)
        for row in rows:
            if not row or len(row) <= idx_employer:
                continue
            status = str(row[idx_status] or "").strip().upper()
            if status not in CERTIFIED_STATUSES:
                continue
            filed = str(row[idx_employer] or "").strip()
            if not filed:
                continue
            canon = canonicalize(filed)
            filed_names[filed.upper()].add(canon)
        wb.close()

    for uscis_path in uscis_paths:
        ingest_uscis_csv(uscis_path, buckets)
        fh, delim = _open_uscis(uscis_path)
        with fh:
            reader = csv.reader(fh, delimiter=delim)
            header = tuple(next(reader, []))
            cols = resolve_uscis_columns(header)
            idx_emp = [h.strip() for h in header].index(cols.employer)
            for row in reader:
                if len(row) <= idx_emp:
                    continue
                filed = row[idx_emp].strip()
                if filed:
                    filed_names[filed.upper()].add(canonicalize(filed))

    write_database(output, buckets, filed_names)


def build_fixture_database(fixtures_dir: Path, output: Path) -> None:
    dol_files = sorted(fixtures_dir.glob("dol_*.xlsx"))
    uscis_files = sorted(fixtures_dir.glob("uscis_*.csv"))
    if not dol_files:
        raise FileNotFoundError(f"No dol_*.xlsx fixtures in {fixtures_dir}")
    build_from_paths(dol_files, uscis_files, output)


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description="Build h1b_data.db from DOL + USCIS files")
    parser.add_argument("--output", type=Path, default=Path("data/h1b_data.db"))
    parser.add_argument("--fixtures", type=Path, help="Directory with dol_*.xlsx and uscis_*.csv")
    parser.add_argument("--dol", type=Path, action="append", default=[], help="DOL xlsx file(s)")
    parser.add_argument("--uscis", type=Path, action="append", default=[], help="USCIS csv file(s)")
    args = parser.parse_args(argv)

    if args.fixtures:
        build_fixture_database(args.fixtures, args.output)
    elif args.dol:
        build_from_paths(args.dol, args.uscis, args.output)
    else:
        parser.error("Provide --fixtures DIR or at least one --dol FILE")

    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
