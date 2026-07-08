#!/usr/bin/env python3
"""Pre-build validation gate: verify USCIS xlsx ingest against an independent
recomputation on real files, plus the FY2020 cross-vintage identity and a
one-quarter DOL join check. Exit 0 = safe to run the full build."""

from __future__ import annotations

import random
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from etl.build import YearBucket, ingest_dol_xlsx, ingest_uscis_xlsx  # noqa: E402
from etl.canonicalize import canonicalize  # noqa: E402

SRC = ROOT / "data" / "sources"
# AMAZON matched by bare prefix: USCIS vintages flip between "AMAZON.COM
# SERVICES" (-> AMAZONCOM SERVICES) and "AMAZON COM SERVICES" (with space) —
# known entity fragmentation, not an ingest bug.
ANCHORS = ["GOOGLE", "AMAZON", "INFOSYS", "DELOITTE CONSULTING", "MICROSOFT"]
# Legacy FY2020 export's Initial Approval total = 121,874 NewEmp + 1,020 NewConc
FY2020_NEW_APPROVALS_TOTAL = 122_894
RANDOM_SAMPLE = 10
DOL_JOIN_FILE = "LCA_Disclosure_Data_FY2025_Q4.xlsx"

failures: list[str] = []


def check(name: str, ok: bool, detail: str = "") -> None:
    print(f"  [{'PASS' if ok else 'FAIL'}] {name}" + (f" — {detail}" if detail else ""))
    if not ok:
        failures.append(name)


blank_name_new_approvals: dict[int, int] = defaultdict(int)


def independent_sums(path: Path) -> dict[tuple[str, int], tuple[int, int, int, int]]:
    """Second code path: raw openpyxl scan, positional column lookup by header
    name, no shared ingest helpers. Returns (new_app, new_den, tr_app, tr_den).
    Side effect: tallies blank-employer-name new approvals per FY (the ingest
    skips those rows; the legacy FY2020 total includes them)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    col = {name: header.index(name) for name in header}

    def num(cell) -> int:
        if cell is None:
            return 0
        return int(str(cell).replace(",", "").strip() or 0)

    sums: dict[tuple[str, int], list[int]] = defaultdict(lambda: [0, 0, 0, 0])
    for row in rows:
        name = str(row[col["Employer (Petitioner) Name"]] or "").strip()
        fy = int(float(str(row[col["Fiscal Year"]]).strip()))
        if not name:
            blank_name_new_approvals[fy] += num(row[col["New Employment Approval"]]) + num(
                row[col["New Concurrent Approval"]]
            )
            continue
        key = (canonicalize(name), fy)
        s = sums[key]
        s[0] += num(row[col["New Employment Approval"]]) + num(
            row[col["New Concurrent Approval"]]
        )
        s[1] += num(row[col["New Employment Denial"]]) + num(row[col["New Concurrent Denial"]])
        s[2] += num(row[col["Change of Employer Approval"]])
        s[3] += num(row[col["Change of Employer Denial"]])
    wb.close()
    return {k: tuple(v) for k, v in sums.items()}


def main() -> int:
    uscis_files = sorted(SRC.glob("Employer Information*.xlsx"))
    if not uscis_files:
        print(f"No USCIS xlsx files in {SRC}")
        return 1

    print("== 1. Ingest vs independent recomputation ==")
    buckets: dict[tuple[str, int], YearBucket] = {}
    expected: dict[tuple[str, int], tuple[int, int, int, int]] = {}
    for f in uscis_files:
        print(f"  reading {f.name} ...")
        ingest_uscis_xlsx(f, buckets)
        expected.update(independent_sums(f))

    rng = random.Random(42)
    sample_keys = [k for k in expected if any(k[0].startswith(a) for a in ANCHORS)]
    sample_keys += rng.sample(sorted(expected.keys()), RANDOM_SAMPLE)
    for key in sample_keys:
        exp = expected[key]
        b = buckets.get(key)
        got = (
            (b.new_approvals, b.new_denials, b.transfer_approvals or 0, b.transfer_denials or 0)
            if b
            else (0, 0, 0, 0)
        )
        check(f"{key[0]} FY{key[1]}", got == exp, f"ingest={got} independent={exp}")

    print("== 2. FY2020 cross-vintage identity ==")
    fy2020_new = sum(b.new_approvals for (_, fy), b in buckets.items() if fy == 2020)
    fy2020_blank = blank_name_new_approvals.get(2020, 0)
    check(
        "FY2020 new_h1b approvals (+ skipped blank-name rows) == legacy Initial total",
        fy2020_new + fy2020_blank == FY2020_NEW_APPROVALS_TOTAL,
        f"ingested {fy2020_new:,} + blank-name {fy2020_blank} vs {FY2020_NEW_APPROVALS_TOTAL:,}",
    )

    print("== 3. Transfer breakout populated ==")
    with_breakout = sum(1 for b in buckets.values() if b.transfer_approvals is not None)
    check(
        "every ingested bucket has a transfer breakout (no NULLs from real files)",
        with_breakout == len(buckets),
        f"{with_breakout:,}/{len(buckets):,}",
    )

    print("== 4. DOL join check (one quarter) ==")
    dol = SRC / DOL_JOIN_FILE
    if dol.exists():
        print(f"  reading {dol.name} (slow, ~1-2 min) ...")
        dol_buckets: dict[tuple[str, int], YearBucket] = defaultdict(YearBucket)
        ingest_dol_xlsx(dol, dol_buckets)
        for anchor in ANCHORS:
            lca = sum(
                b.certified for (name, _), b in dol_buckets.items() if name.startswith(anchor)
            )
            uscis = sum(
                b.new_approvals
                for (name, fy), b in buckets.items()
                if fy == 2025 and name.startswith(anchor)
            )
            check(
                f"{anchor}: LCA and USCIS on same canonical key",
                lca > 0 and uscis > 0,
                f"lca={lca:,} uscis_new={uscis:,}",
            )
    else:
        check(f"{DOL_JOIN_FILE} present", False, "file missing")

    print()
    if failures:
        print(f"GATE CLOSED — {len(failures)} failure(s): {failures}")
        return 1
    print("GATE OPEN — safe to run the full build.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
